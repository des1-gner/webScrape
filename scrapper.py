import xml.etree.ElementTree as ET

import openpyxl

import os

def process_xml(file_paths, output_file):

    # Create a new workbook and select the active sheet

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Initialize the row counter

    row_num = 1

    for file_path in file_paths:

        tree = ET.parse(file_path)
        root = tree.getroot()

        # Iterate over each <item> element

        for item in root.findall('./channel/item'):

            # Extract the desired tags

            title = item.find('title').text if item.find('title') is not None else ""

            post_id = item.find('ns0:post_id', {'ns0': 'http://wordpress.org/export/1.2/'}).text if item.find('ns0:post_id', {'ns0': 'http://wordpress.org/export/1.2/'}) is not None else ""
            
            link = item.find('link').text if item.find('link') is not None else ""
            
            pub_date = item.find('pubDate').text if item.find('pubDate') is not None else ""
            
            post_modified = item.find('ns0:post_modified', {'ns0': 'http://wordpress.org/export/1.2/'}).text if item.find('ns0:post_modified', {'ns0': 'http://wordpress.org/export/1.2/'}) is not None else ""
            
            encoded = item.find('ns2:encoded', {'ns2': 'http://purl.org/rss/1.0/modules/content/'}).text if item.find('ns2:encoded', {'ns2': 'http://purl.org/rss/1.0/modules/content/'}) is not None else ""
            
            encoded_excerpt = item.find('ns3:encoded', {'ns3': 'http://wordpress.org/export/1.2/excerpt/'}).text if item.find('ns3:encoded', {'ns3': 'http://wordpress.org/export/1.2/excerpt/'}) is not None else ""
            
            # Skip adding the row if the "Content" part is empty or None

            if not encoded or not encoded.strip():

                continue

            # Combine all the content into a single string

            combined_content = f"Post ID: {post_id}\nLink: {link}\nPublication Date: {pub_date}\nUpdated Date: {post_modified}\n\nContent:\n{encoded}\n\nExcerpt:\n{encoded_excerpt}"

            # Write the extracted data to the sheet

            sheet.cell(row=row_num, column=1, value=title)
            sheet.cell(row=row_num, column=2, value=combined_content)

            # Increment the row counter

            row_num += 1

    # Save the workbook to the output file

    workbook.save(output_file)

# Raw Wordpress Export Files

xml_file_paths = [

    'wpExports/ibuildPosts.xml',
    'wpExports/ibuildKnowledgeBase.xml',
    'wpExports/ibuildLandingPage.xml',
    'wpExports/ibuildProperties.xml',
    'wpExports/ibuildPages.xml'

]

output_file_path = 'ibuildPosts_combined.xlsx'

process_xml(xml_file_paths, output_file_path)